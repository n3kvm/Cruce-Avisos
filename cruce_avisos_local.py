import argparse
import os
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
PDF_EXTENSIONS = {".pdf"}


def normalize_cotizacion(value):
    text = str(value or "").upper()
    text = re.sub(r"\.(XLSX|XLSM|XLS|PDF)$", "", text)
    text = re.sub(r"\s*\([0-9]+\)\s*$", "", text)
    text = re.sub(r"\s*-\s*COPIA\s*$", "", text)
    text = re.sub(r"\s*-\s*FUSIONADO\s*$", "", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def normalize_aviso(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    digits = re.sub(r"\D+", "", text)
    return digits.strip()


def read_master(master_path):
    df = pd.read_excel(master_path, dtype=str)
    aviso_cols = [c for c in df.columns if str(c).strip().casefold() == "aviso"]
    if not aviso_cols:
        raise ValueError(f"No encontre una columna llamada 'Aviso'. Columnas: {list(df.columns)}")
    aviso_col = aviso_cols[0]
    df["Aviso_normalizado"] = df[aviso_col].map(normalize_aviso)
    df = df[df["Aviso_normalizado"] != ""].copy()
    return df, aviso_col


def list_local_files(folder):
    paths = []
    for path in folder.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        suffix = path.suffix.lower()
        if suffix in EXCEL_EXTENSIONS or suffix in PDF_EXTENSIONS:
            paths.append(path)
    return sorted(paths, key=lambda p: str(p).casefold())


def scan_filename(path, aviso_set):
    hits = []
    for candidate in set(re.findall(r"\d{6,12}", path.name)):
        if candidate in aviso_set:
            hits.append({
                "Aviso": candidate,
                "Archivo": path.name,
                "Ruta": str(path),
                "Tipo hallazgo": "Nombre archivo",
                "Hoja": "",
                "Celda": "",
                "Valor celda": path.name,
                "Modificado": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    return hits


def scan_workbook(path, aviso_set):
    hits = []
    errors = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return hits, [{"Archivo": path.name, "Ruta": str(path), "Hoja": "", "Error": str(exc)}]

    token_regex = re.compile(r"\d{6,12}")
    for sheet in wb.worksheets:
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    candidates = set()
                    normalized = normalize_aviso(value)
                    if normalized:
                        candidates.add(normalized)
                    if isinstance(value, str):
                        candidates.update(token_regex.findall(value))
                    for candidate in candidates.intersection(aviso_set):
                        hits.append({
                            "Aviso": candidate,
                            "Archivo": path.name,
                            "Ruta": str(path),
                            "Tipo hallazgo": "Contenido Excel",
                            "Hoja": sheet.title,
                            "Celda": cell.coordinate,
                            "Valor celda": str(value)[:250],
                            "Modificado": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                        })
        except Exception as exc:
            errors.append({"Archivo": path.name, "Ruta": str(path), "Hoja": sheet.title, "Error": str(exc)})
    return hits, errors


def build_dashboard(master_df, files_meta, hits, errors, output_path):
    hits_df = pd.DataFrame(hits)
    if hits_df.empty:
        hits_df = pd.DataFrame(columns=["Aviso", "Archivo", "Ruta", "Tipo hallazgo", "Hoja", "Celda", "Valor celda", "Modificado"])

    hits_df["Cotizacion_key"] = hits_df["Archivo"].map(normalize_cotizacion)
    grouped = hits_df.groupby("Aviso", dropna=False).agg(
        Hallazgos=("Archivo", "count"),
        Veces_encontrado=("Cotizacion_key", lambda x: len(set(v for v in x if v))),
        Archivos=("Archivo", lambda x: " | ".join(sorted(set(map(str, x))))),
        Tipos_hallazgo=("Tipo hallazgo", lambda x: " | ".join(sorted(set(map(str, x))))),
        Hojas=("Hoja", lambda x: " | ".join(sorted(set(v for v in map(str, x) if v)))),
    ).reset_index()

    result = master_df.merge(grouped, left_on="Aviso_normalizado", right_on="Aviso", how="left")
    result["Veces_encontrado"] = result["Veces_encontrado"].fillna(0).astype(int)
    result["Hallazgos"] = result["Hallazgos"].fillna(0).astype(int)
    result["Estado montaje"] = result["Veces_encontrado"].map(
        lambda n: "Pendiente" if n == 0 else ("Montado" if n == 1 else "Duplicado")
    )
    result["PTBO"] = result["Status sistema"].fillna("").str.contains("PTBO", case=False, regex=False)
    for col in ["Archivos", "Tipos_hallazgo", "Hojas"]:
        result[col] = result[col].fillna("")

    summary = pd.DataFrame([
        {"Indicador": "Total avisos maestro", "Valor": len(result)},
        {"Indicador": "Montados", "Valor": int((result["Estado montaje"] == "Montado").sum())},
        {"Indicador": "Pendientes", "Valor": int((result["Estado montaje"] == "Pendiente").sum())},
        {"Indicador": "Duplicados", "Valor": int((result["Estado montaje"] == "Duplicado").sum())},
        {"Indicador": "PTBO / peticion de borrado", "Valor": int(result["PTBO"].sum())},
        {"Indicador": "Archivos revisados", "Valor": len(files_meta)},
        {"Indicador": "Hallazgos", "Valor": len(hits_df)},
        {"Indicador": "Errores de lectura", "Valor": len(errors)},
        {"Indicador": "Fecha ejecucion", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ])

    files_df = pd.DataFrame(files_meta)
    errors_df = pd.DataFrame(errors) if errors else pd.DataFrame(columns=["Archivo", "Ruta", "Hoja", "Error"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Dashboard", index=False)
        result.to_excel(writer, sheet_name="Avisos_estado", index=False)
        hits_df.to_excel(writer, sheet_name="Hallazgos", index=False)
        files_df.to_excel(writer, sheet_name="Archivos_revisados", index=False)
        errors_df.to_excel(writer, sheet_name="Errores", index=False)

    wb = openpyxl.load_workbook(output_path)
    colors = {
        "Pendiente": "F4CCCC",
        "Montado": "D9EAD3",
        "Duplicado": "FCE5CD",
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        if ws.title == "Avisos_estado":
            headers = {cell.value: cell.column for cell in ws[1]}
            status_col = headers.get("Estado montaje")
            if status_col:
                for row in range(2, ws.max_row + 1):
                    status = ws.cell(row, status_col).value
                    fill = colors.get(status)
                    if fill:
                        ws.cell(row, status_col).fill = openpyxl.styles.PatternFill("solid", fgColor=fill)
        for column_cells in ws.columns:
            values = [str(c.value) for c in column_cells if c.value is not None]
            width = min(max([len(v) for v in values] + [10]) + 2, 70)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Cruza avisos maestro contra archivos locales sincronizados desde Teams.")
    base_dir = Path(__file__).resolve().parent
    parser.add_argument("--master", default=os.getenv("AVISOS_MASTER_PATH", str(base_dir / "data" / "AVISOS.xlsx")))
    parser.add_argument("--folder", default=os.getenv("SOPORTES_FOLDER_PATH", str(base_dir / "soportes")))
    parser.add_argument("--output", default=os.getenv("CRUCE_OUTPUT_PATH", str(base_dir / "salidas" / "dashboard_avisos_local.xlsx")))
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists():
        raise FileNotFoundError(f"No existe la carpeta: {folder}")

    print(f"Leyendo maestro: {args.master}")
    master_df, aviso_col = read_master(args.master)
    aviso_set = set(master_df["Aviso_normalizado"])
    print(f"Avisos cargados: {len(aviso_set)} | columna: {aviso_col}")

    files = list_local_files(folder)
    print(f"Archivos locales encontrados: {len(files)}")

    all_hits = []
    all_errors = []
    files_meta = []
    for index, path in enumerate(files, start=1):
        print(f"[{index}/{len(files)}] Revisando: {path.name}")
        stat = path.stat()
        files_meta.append({
            "Archivo": path.name,
            "Ruta": str(path),
            "Extension": path.suffix.lower(),
            "Tamano bytes": stat.st_size,
            "Modificado": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        })
        all_hits.extend(scan_filename(path, aviso_set))
        if path.suffix.lower() in EXCEL_EXTENSIONS:
            hits, errors = scan_workbook(path, aviso_set)
            all_hits.extend(hits)
            all_errors.extend(errors)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    build_dashboard(master_df, files_meta, all_hits, all_errors, output_path)
    print(f"Dashboard generado: {output_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        traceback.print_exc()
        sys.exit(1)
