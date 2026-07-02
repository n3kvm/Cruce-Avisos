import base64
import json
import mimetypes
import os
import subprocess
import sys
import re
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
TOKEN_CACHE = BASE_DIR / ".graph_backend_token.json"
FILES_CACHE = BASE_DIR / "backend_graph_cache.json"
DASHBOARD_DIR = BASE_DIR / "dashboard_html"
GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
HOST = os.getenv("BACKEND_HOST", "0.0.0.0")
PORT = int(os.getenv("BACKEND_PORT", "8765"))
CLIENT_ID = os.getenv("CLIENT_ID", "e90f6053-e70c-409c-9dcf-1d097f86a4a5")
TENANT_ID = os.getenv("TENANT_ID", "organizations")
SITE_HOST = os.getenv("SITE_HOST", "brillaseo2.sharepoint.com")
SITE_PATH = os.getenv("SITE_PATH", "/sites/SoportesEspejo")
DRIVE_NAME = os.getenv("DRIVE_NAME", "Shared Documents")
FOLDER_PATH = os.getenv("FOLDER_PATH", "")
SCOPES = os.getenv("GRAPH_SCOPES", "https://graph.microsoft.com/Files.Read.All offline_access").split()
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}



def static_response(handler, relative_path):
    requested = "index.html" if relative_path in {"", "/"} else relative_path.lstrip("/")
    requested = urllib.parse.unquote(requested).replace("\\", "/")
    dashboard_root = DASHBOARD_DIR.resolve()
    target = (dashboard_root / requested).resolve()
    try:
        target.relative_to(dashboard_root)
    except ValueError:
        json_response(handler, 403, {"ok": False, "error": "Ruta no permitida"})
        return
    if not target.is_file():
        json_response(handler, 404, {"ok": False, "error": f"Archivo no encontrado: {requested}", "root": str(dashboard_root)})
        return
    content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
    raw = target.read_bytes()
    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(raw)))
    handler.end_headers()
    handler.wfile.write(raw)

def json_response(handler, status, payload):
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    handler.send_header("Content-Length", str(len(raw)))
    handler.end_headers()
    handler.wfile.write(raw)


def normalize_header(value):
    import unicodedata
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return text


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", "", normalize_header(value))


def normalize_aviso(value):
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\D+", "", text).strip()


def normalize_cotizacion(value):
    import unicodedata
    text = str(value or "").upper()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\.(XLSX|XLSM|XLS|PDF)$", "", text, flags=re.I)
    text = re.sub(r"\s*\([0-9]+\)\s*$", "", text)
    text = re.sub(r"\s*-\s*COPIA\s*$", "", text)
    text = re.sub(r"\s*-\s*FUSIONADO\s*$", "", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def clean(value):
    return str(value or "").strip()


def request_form(url, data):
    encoded = urllib.parse.urlencode(data).encode("utf-8")
    req = urllib.request.Request(url, data=encoded, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def graph_request(path_or_url, token, *, method="GET", body=None, content_type="application/json"):
    url = path_or_url if str(path_or_url).startswith("http") else GRAPH_ROOT + path_or_url
    data = None if body is None else (body if isinstance(body, bytes) else json.dumps(body).encode("utf-8"))
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    if body is not None:
        req.add_header("Content-Type", content_type)
    try:
        with urllib.request.urlopen(req, timeout=180) as response:
            raw = response.read()
            if not raw:
                return None
            content_type = response.headers.get("Content-Type", "")
            if "application/json" in content_type:
                return json.loads(raw.decode("utf-8"))
            return raw
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Graph error {exc.code}: {detail}") from exc


def load_cached_token():
    if not TOKEN_CACHE.exists():
        return None
    try:
        cache = json.loads(TOKEN_CACHE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if cache.get("expires_at", 0) - 180 > time.time():
        return cache.get("access_token")
    refresh = cache.get("refresh_token")
    if not refresh:
        return None
    try:
        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
        token = request_form(token_url, {
            "grant_type": "refresh_token",
            "client_id": CLIENT_ID,
            "refresh_token": refresh,
            "scope": " ".join(SCOPES),
        })
        token["expires_at"] = time.time() + int(token.get("expires_in", 3600))
        TOKEN_CACHE.write_text(json.dumps(token, indent=2), encoding="utf-8")
        return token["access_token"]
    except Exception:
        return None


def start_device_flow():
    cached = load_cached_token()
    if cached:
        return {"authenticated": True, "message": "Sesion Graph activa."}
    device_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/devicecode"
    device = request_form(device_url, {"client_id": CLIENT_ID, "scope": " ".join(SCOPES)})
    DEVICE_FLOW.clear()
    DEVICE_FLOW.update({
        "device_code": device["device_code"],
        "user_code": device.get("user_code"),
        "verification_uri": device.get("verification_uri"),
        "verification_uri_complete": device.get("verification_uri_complete"),
        "message": device.get("message"),
        "interval": int(device.get("interval", 5)),
        "expires_at": time.time() + int(device.get("expires_in", 900)),
        "last_poll": 0,
    })
    return {k: DEVICE_FLOW.get(k) for k in ["user_code", "verification_uri", "verification_uri_complete", "message", "expires_at"]} | {"authenticated": False}


def poll_device_flow():
    cached = load_cached_token()
    if cached:
        return {"authenticated": True, "message": "Sesion Graph activa."}
    if not DEVICE_FLOW:
        return {"authenticated": False, "needs_start": True, "message": "No hay autenticacion iniciada."}
    if time.time() > DEVICE_FLOW.get("expires_at", 0):
        DEVICE_FLOW.clear()
        return {"authenticated": False, "expired": True, "message": "El codigo expiro. Genera uno nuevo."}
    if time.time() - DEVICE_FLOW.get("last_poll", 0) < DEVICE_FLOW.get("interval", 5):
        return {"authenticated": False, "pending": True, "message": "Esperando autorizacion Microsoft."}
    DEVICE_FLOW["last_poll"] = time.time()
    token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    try:
        token = request_form(token_url, {
            "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
            "client_id": CLIENT_ID,
            "device_code": DEVICE_FLOW["device_code"],
        })
        token["expires_at"] = time.time() + int(token.get("expires_in", 3600))
        TOKEN_CACHE.write_text(json.dumps(token, indent=2), encoding="utf-8")
        DEVICE_FLOW.clear()
        return {"authenticated": True, "message": "Autenticacion completada."}
    except urllib.error.HTTPError as exc:
        payload = json.loads(exc.read().decode("utf-8"))
        error = payload.get("error")
        if error in {"authorization_pending", "slow_down"}:
            if error == "slow_down":
                DEVICE_FLOW["interval"] = DEVICE_FLOW.get("interval", 5) + 5
            return {"authenticated": False, "pending": True, "message": "Esperando autorizacion Microsoft."}
        DEVICE_FLOW.clear()
        return {"authenticated": False, "error": payload, "message": str(payload)}


def get_token():
    cached = load_cached_token()
    if cached:
        return cached
    raise PermissionError("AUTH_REQUIRED")


def paged(path, token):
    out = []
    url = path
    while url:
        payload = graph_request(url, token)
        out.extend(payload.get("value", []))
        url = payload.get("@odata.nextLink")
    return out


def get_drive_and_folder(token):
    site = graph_request(f"/sites/{SITE_HOST}:{SITE_PATH}", token)
    drives = paged(f"/sites/{site['id']}/drives?$select=id,name,webUrl", token)
    wanted = {DRIVE_NAME.casefold(), "documentos compartidos", "shared documents", "documents", "documentos"}
    drive = next((d for d in drives if d.get("name", "").casefold() in wanted), drives[0] if drives else None)
    if not drive:
        raise RuntimeError("No encontre biblioteca de documentos en SharePoint.")
    folder_path = FOLDER_PATH.strip("/")
    if folder_path:
        folder_encoded = urllib.parse.quote(folder_path)
        folder = graph_request(f"/drives/{drive['id']}/root:/{folder_encoded}", token)
    else:
        folder = graph_request(f"/drives/{drive['id']}/root", token)
    return drive, folder


def list_excel_files(token):
    drive, folder = get_drive_and_folder(token)
    pending = [folder["id"]]
    files = []
    while pending:
        item_id = pending.pop()
        children = paged(f"/drives/{drive['id']}/items/{item_id}/children?$select=id,name,folder,file,lastModifiedDateTime,webUrl,size,eTag", token)
        for item in children:
            if item.get("folder"):
                pending.append(item["id"])
            elif Path(item.get("name", "")).suffix.lower() in EXCEL_EXTENSIONS and not item.get("name", "").startswith("~$"):
                item["driveId"] = drive["id"]
                files.append(item)
    return files


def download_item(token, item):
    return graph_request(f"/drives/{item['driveId']}/items/{item['id']}/content", token)


def scan_workbook_bytes(content, item, aviso_set):
    hits = []
    errors = []
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return hits, [{"archivo": item.get("name"), "error": str(exc)}]
    token_regex = re.compile(r"\d{6,12}")
    for sheet in wb.worksheets:
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    candidates = set()
                    normalized = normalize_aviso(value)
                    if normalized:
                        candidates.add(normalized)
                    if isinstance(value, str):
                        candidates.update(token_regex.findall(value))
                    for candidate in candidates.intersection(aviso_set):
                        hits.append({
                            "aviso": candidate,
                            "archivo": item.get("name"),
                            "tipo": "Contenido Excel",
                            "hoja": sheet.title,
                            "celda": cell.coordinate,
                            "valor": str(value)[:250],
                            "webUrl": item.get("webUrl", ""),
                            "cotizacionKey": normalize_cotizacion(item.get("name")),
                        })
        except Exception as exc:
            errors.append({"archivo": item.get("name"), "hoja": sheet.title, "error": str(exc)})
    for candidate in set(token_regex.findall(item.get("name", ""))).intersection(aviso_set):
        hits.append({
            "aviso": candidate,
            "archivo": item.get("name"),
            "tipo": "Nombre archivo",
            "hoja": "",
            "celda": "",
            "valor": item.get("name"),
            "webUrl": item.get("webUrl", ""),
            "cotizacionKey": normalize_cotizacion(item.get("name")),
        })
    return hits, errors


def load_cache():
    if not FILES_CACHE.exists():
        return {"files": {}}
    try:
        return json.loads(FILES_CACHE.read_text(encoding="utf-8"))
    except Exception:
        return {"files": {}}


def save_cache(cache):
    FILES_CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")


def get_hallazgos_for_avisos(aviso_set):
    token = get_token()
    cache = load_cache()
    file_cache = cache.setdefault("files", {})
    files = list_excel_files(token)
    all_hits = []
    errors = []
    scanned = 0
    reused = 0
    for item in files:
        cache_key = item["id"]
        signature = f"{item.get('eTag')}|{item.get('lastModifiedDateTime')}|{item.get('size')}"
        cached = file_cache.get(cache_key)
        if cached and cached.get("signature") == signature:
            hits = [h for h in cached.get("hits", []) if h.get("aviso") in aviso_set]
            all_hits.extend(hits)
            reused += 1
            continue
        content = download_item(token, item)
        hits, file_errors = scan_workbook_bytes(content, item, aviso_set)
        file_cache[cache_key] = {"signature": signature, "name": item.get("name"), "hits": hits, "errors": file_errors}
        all_hits.extend(hits)
        errors.extend(file_errors)
        scanned += 1
    save_cache(cache)
    return all_hits, errors, {"archivos": len(files), "escaneados": scanned, "cache": reused}


def run_local_refresh():
    scripts = [
        BASE_DIR / "cruce_avisos_local.py",
        BASE_DIR / "powerbi" / "generar_powerbi_data.py",
        DASHBOARD_DIR / "regenerar_data_corregida.py",
    ]
    logs = []
    started = datetime.now()
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    for script in scripts:
        if not script.exists():
            raise FileNotFoundError(f"No existe el script requerido: {script}")
        proc = subprocess.run(
            [sys.executable, str(script)],
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=600,
            env=env,
        )
        logs.append({
            "script": script.name,
            "ok": proc.returncode == 0,
            "returncode": proc.returncode,
            "stdout": proc.stdout[-4000:],
            "stderr": proc.stderr[-4000:],
        })
        if proc.returncode != 0:
            raise RuntimeError(f"Fallo {script.name}: {proc.stderr or proc.stdout}")
    return {
        "ok": True,
        "started": started.strftime("%Y-%m-%d %H:%M:%S"),
        "finished": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "logs": logs,
    }

def parse_uploaded_file(content_type, body):
    match = re.search(r"boundary=([^;]+)", content_type or "")
    if not match:
        raise ValueError("La solicitud no trae multipart/form-data.")
    boundary = ("--" + match.group(1).strip().strip('"')).encode("utf-8")
    for part in body.split(boundary):
        if b"Content-Disposition" not in part:
            continue
        header, _, payload = part.partition(b"\r\n\r\n")
        if b'name="file"' not in header:
            continue
        filename_match = re.search(rb'filename="([^"]+)"', header)
        filename = filename_match.group(1).decode("utf-8", errors="replace") if filename_match else "AVISOS.xlsx"
        payload = payload.rstrip(b"\r\n-")
        return filename, payload
    raise ValueError("No encontre el archivo XLSX en la solicitud.")


def read_master_from_bytes(content):
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(v) for v in next(rows)]
    header_map = {normalize_key(h): idx for idx, h in enumerate(headers)}
    aviso_idx = header_map.get("aviso")
    if aviso_idx is None:
        raise ValueError(f"No encontre una columna llamada Aviso. Columnas: {headers}")

    def row_value(row, *names):
        for name in names:
            idx = header_map.get(normalize_key(name))
            if idx is not None and idx < len(row):
                return row[idx]
        return ""

    avisos = []
    for row in rows:
        aviso = normalize_aviso(row[aviso_idx] if aviso_idx < len(row) else "")
        if not aviso:
            continue
        fecha = format_date(row_value(row, "Fecha de aviso", "Fecha"))
        status = clean(row_value(row, "Status sistema"))
        avisos.append({
            "aviso": aviso,
            "fecha": fecha,
            "anio": fecha[:4] if fecha else "",
            "mes": fecha[5:7] if len(fecha) >= 7 else "",
            "grupo": clean(row_value(row, "Grupo planif.", "Grupo")),
            "emplazamiento": clean(row_value(row, "Emplazamiento")),
            "contratista": clean(row_value(row, "Nombre Contratista", "Contratista")),
            "prioridad": clean(row_value(row, "Prioridad")),
            "statusSistema": status,
            "codif": clean(row_value(row, "Codif.txt.cod.", "Codif.txt.cód.", "Codif")),
            "ptbo": "PTBO" in status.upper(),
            "statusUsuario": clean(row_value(row, "Status usuario")),
            "descripcion": clean(row_value(row, "Descripcion", "Descripción")),
            "denominacion": clean(row_value(row, "Denominacion", "Denominación")),
        })
    return avisos


def format_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text.replace("Z", ""))
        return parsed.strftime("%Y-%m-%d")
    except Exception:
        pass
    return text[:10]


def build_dataset(master_avisos, hallazgos, stats):
    by_aviso = defaultdict(list)
    for hit in hallazgos:
        by_aviso[hit["aviso"]].append(hit)
    avisos = []
    for item in master_avisos:
        hits = by_aviso.get(item["aviso"], [])
        quote_keys = sorted({h.get("cotizacionKey") or normalize_cotizacion(h.get("archivo")) for h in hits if h.get("archivo")})
        files = sorted({h.get("archivo") for h in hits if h.get("archivo")})
        types = sorted({h.get("tipo") for h in hits if h.get("tipo")})
        veces = len([q for q in quote_keys if q])
        aviso = dict(item)
        aviso.update({
            "estado": "Pendiente" if veces == 0 else ("Montado" if veces == 1 else "Duplicado"),
            "veces": veces,
            "hallazgos": len(hits),
            "archivos": " | ".join(files),
            "tipos": " | ".join(types),
        })
        avisos.append(aviso)
    estados = sorted({a["estado"] for a in avisos})
    meta = {
        "totalAvisos": len(avisos),
        "montados": sum(1 for a in avisos if a["estado"] == "Montado"),
        "pendientes": sum(1 for a in avisos if a["estado"] == "Pendiente"),
        "duplicados": sum(1 for a in avisos if a["estado"] == "Duplicado"),
        "hallazgos": len(hallazgos),
        "archivos": stats.get("archivos", 0),
        "estados": estados,
        "ptbo": sum(1 for a in avisos if a.get("ptbo")),
        "graph": stats,
    }
    return {"meta": meta, "avisos": avisos, "hallazgos": hallazgos}


class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        json_response(self, 200, {"ok": True})

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/status"):
            json_response(self, 200, {"ok": True, "mode": "graph", "site": f"{SITE_HOST}{SITE_PATH}", "folder": FOLDER_PATH})
            return
        if path.startswith("/api/"):
            json_response(self, 404, {"ok": False, "error": "Ruta no encontrada"})
            return
        static_response(self, path)

    def do_POST(self):
        path = urllib.parse.urlparse(self.path).path
        if path.startswith("/api/refresh"):
            try:
                result = run_local_refresh()
                json_response(self, 200, result)
            except Exception as exc:
                json_response(self, 500, {"ok": False, "error": str(exc)})
            return
        if not path.startswith("/api/cruce"):
            json_response(self, 404, {"ok": False, "error": "Ruta no encontrada"})
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            filename, content = parse_uploaded_file(self.headers.get("Content-Type", ""), body)
            master_avisos = read_master_from_bytes(content)
            aviso_set = {a["aviso"] for a in master_avisos}
            hallazgos, errors, stats = get_hallazgos_for_avisos(aviso_set)
            dataset = build_dataset(master_avisos, hallazgos, stats)
            dataset["source"] = filename
            dataset["errors"] = errors[:50]
            json_response(self, 200, {"ok": True, "data": dataset})
        except Exception as exc:
            json_response(self, 500, {"ok": False, "error": str(exc)})

    def log_message(self, fmt, *args):
        print(f"{self.address_string()} - {fmt % args}")


if __name__ == "__main__":
    print(f"Backend Graph escuchando en http://{HOST}:{PORT}")
    print(f"SharePoint: {SITE_HOST}{SITE_PATH} / {FOLDER_PATH}")
    print("Deja esta ventana abierta mientras otros consultan el dashboard.")
    ThreadingHTTPServer((HOST, PORT), Handler).serve_forever()





